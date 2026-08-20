#!/usr/bin/env python3
"""Konversi file master SWS (.xlsx) menjadi data/dashboard.json untuk dashboard publik.

Pemakaian:
    python scripts/convert_sws.py data/raw/SWS_W33__G2.xlsx
    python scripts/convert_sws.py data/raw/SWS_W33__G2.xlsx -o data/dashboard.json

Sheet yang dibaca:
    EXT  -> perpanjangan kontrak yang sudah dilakukan AFPS
    NOO  -> outlet NOO yang sudah deal / sudah terbit nomor AP
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TAHUN_VALID = (2025, 2026)


# ---------------------------------------------------------------- util

def norm(v) -> str:
    """Rapikan teks sel: buang newline, spasi ganda, jadikan UPPERCASE."""
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    return "" if s.lower() in ("none", "-", "#n/a", "n/a") else s


def upper(v) -> str:
    return norm(v).upper()


BULAN_ID = {
    "jan": 1, "feb": 2, "peb": 2, "mar": 3, "mrt": 3, "apr": 4, "mei": 5, "may": 5,
    "jun": 6, "jul": 7, "agu": 8, "ags": 8, "agt": 8, "aug": 8, "sep": 9, "okt": 10,
    "oct": 10, "nov": 11, "nop": 11, "des": 12, "dec": 12,
}


def _tahun(y):
    y = int(y)
    return y + 2000 if y < 100 else y


def as_date(v):
    """Tanggal dari sel SWS: objek tanggal, serial Excel, atau teks (termasuk nama bulan Indonesia)."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if 20000 <= v <= 80000:  # serial Excel (basis 1899-12-30)
            return dt.date(1899, 12, 30) + dt.timedelta(days=int(v))
        return None
    s = norm(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, bl, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = re.match(r"^(\d{1,2})[-/. ](\d{1,2})[-/. ](\d{2,4})", s)
        if m:  # urutan Indonesia: hari-bulan-tahun
            d, bl, y = int(m.group(1)), int(m.group(2)), _tahun(m.group(3))
        else:
            m = re.match(r"^(\d{1,2})[-/. ]([A-Za-z]{3,12})[-/. ](\d{2,4})", s)
            if not m:
                return None
            bl = BULAN_ID.get(m.group(2)[:3].lower())
            if not bl:
                return None
            d, y = int(m.group(1)), _tahun(m.group(3))
    try:
        return dt.date(y, bl, d)
    except ValueError:
        return None


def iso(d):
    return d.isoformat() if d else None


def as_num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 2)
    s = norm(v).replace("Rp", "").replace(".", "").replace(",", ".").replace(" ", "")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def as_int(v):
    n = as_num(v)
    return int(n) if n is not None else None


def week_iso(d):
    """Nomor minggu ISO dari tanggal deal (kolom WEEK DEAL di SWS kerap tidak konsisten)."""
    return d.isocalendar()[1] if d else None


def kuartal(d):
    return f"{d.year}-Q{(d.month - 1) // 3 + 1}" if d else None


def cari_baris_header(ws, kata_kunci="NO OUTLET", maks=12):
    """Cari indeks baris header; layout SWS punya beberapa baris judul di atas."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=maks, values_only=True), 1):
        if any(upper(c) == kata_kunci for c in row):
            return i
    raise SystemExit(f"Header '{kata_kunci}' tidak ditemukan di sheet {ws.title}")


class Kolom:
    """Pencari kolom berdasarkan nama header (tahan terhadap newline & spasi ganda).

    Header SWS punya nama ganda (mis. 'NOMOR AP' muncul untuk kontrak lama dan
    kontrak baru), jadi pemilihan bisa memakai occurrence ke-n.
    """

    def __init__(self, header):
        self.peta = defaultdict(list)
        for i, h in enumerate(header):
            key = upper(h)
            if key:
                self.peta[key].append(i)

    def idx(self, nama, ke=0):
        key = upper(nama)
        kandidat = self.peta.get(key, [])
        if not kandidat:  # toleransi perubahan redaksi header: cocokkan awalan
            kandidat = [i for k, v in self.peta.items() if k.startswith(key) for i in v]
            kandidat.sort()
        if not kandidat:
            return None
        try:
            return kandidat[ke]
        except IndexError:
            return kandidat[-1]

    def get(self, row, nama, ke=0):
        i = self.idx(nama, ke)
        return row[i] if i is not None and i < len(row) else None


# ---------------------------------------------------------------- parsing

def baca_sheet(wb, nama_sheet):
    ws = wb[nama_sheet]
    baris_header = cari_baris_header(ws)
    rows = list(ws.iter_rows(min_row=baris_header, values_only=True))
    kol = Kolom(rows[0])
    data = [r for r in rows[1:] if norm(kol.get(r, "No Outlet")) or norm(kol.get(r, "NAMA OUTLET"))]
    return kol, data


def region_per_afps(sumber_baris):
    """Tentukan region & GRFPM tiap AFPS berdasarkan kemunculan terbanyak."""
    hitung = defaultdict(Counter)
    grfpm = defaultdict(Counter)
    for afps, reg, gr in sumber_baris:
        if not afps:
            continue
        if reg:
            hitung[afps][reg] += 1
        if gr:
            grfpm[afps][gr] += 1
    return {
        a: {
            "region": hitung[a].most_common(1)[0][0] if hitung[a] else "TANPA REGION",
            "grfpm": grfpm[a].most_common(1)[0][0] if grfpm[a] else "",
        }
        for a in hitung
    }


def rapikan_grfpm(v):
    s = upper(v)
    m = re.search(r"(\d+)", s)
    return f"GRFPM {int(m.group(1))}" if m and "GRFPM" in s.replace("GRPFM", "GRFPM") else (s or "")


def parse_ext(kol, data, channel_ke_kpi):
    """Perpanjangan kontrak: baris EXT dengan status DEAL."""
    deals, issues, mentah = [], [], []
    for r in data:
        afps = upper(kol.get(r, "Nama AFPS"))
        region = upper(kol.get(r, "REGION"))
        gr = rapikan_grfpm(kol.get(r, "GRFPM"))
        mentah.append((afps, region, gr))

        status = upper(kol.get(r, "DEAL/NO DEAL/ PROSES")) or "BELUM DIVISIT"
        if status.startswith("NO DEAL"):
            status = "NO DEAL"
        channel = upper(kol.get(r, "CHANNEL"))
        if status != "DEAL":
            continue

        no_outlet = norm(kol.get(r, "No Outlet"))
        outlet = norm(kol.get(r, "NAMA OUTLET"))
        tgl_deal = as_date(kol.get(r, "TANGGAL DEAL"))
        ap_baru = norm(kol.get(r, "NOMOR AP", ke=1))
        masalah = []
        if tgl_deal is None:
            masalah.append("Status DEAL tetapi TANGGAL DEAL kosong")
        elif tgl_deal.year not in TAHUN_VALID:
            masalah.append(f"Tanggal deal di luar rentang wajar: {tgl_deal.isoformat()}")
        if not ap_baru:
            masalah.append("Sudah DEAL tetapi nomor AP baru belum terisi")
        for m in masalah:
            issues.append({"sheet": "EXT", "no_outlet": no_outlet, "outlet": outlet,
                           "afps": afps, "region": region, "masalah": m})

        valid = tgl_deal is not None and tgl_deal.year in TAHUN_VALID
        deals.append({
            "sumber": "PERPANJANGAN",
            "no_outlet": no_outlet,
            "outlet": outlet,
            "grfpm": gr,
            "region": region,
            "kota": upper(kol.get(r, "KOTA")),
            "kecamatan": upper(kol.get(r, "KECAMATAN")),
            "afps": afps,
            "jenis": upper(kol.get(r, "JENIS OUTLET")),
            "channel": channel,
            "kpi": channel_ke_kpi.get(channel, "lainnya"),
            "kategori": "",
            "brand": upper(kol.get(r, "BRAND", ke=1)) or upper(kol.get(r, "BRAND")),
            "kontrak": upper(kol.get(r, "BB/BR/BL", ke=1)) or upper(kol.get(r, "BB/BR/BL")),
            "takeover": upper(kol.get(r, "TAKEOVER OR BUKAN TAKEOVER")),
            "alamat": norm(kol.get(r, "ALAMAT OUTLET")),
            "pic": norm(kol.get(r, "NAMA PIC")),
            "telp": norm(kol.get(r, "NO TELP PIC")),
            "kode_outlet": norm(kol.get(r, "KODE OUTLET")),
            "ap": ap_baru,
            "ap_lama": norm(kol.get(r, "NOMOR AP")),
            "tgl_visit": iso(as_date(kol.get(r, "TANGGAL VISIT"))),
            "tgl_deal": iso(tgl_deal),
            "week_deal": as_int(kol.get(r, "WEEK DEAL")),
            "week_iso": week_iso(tgl_deal) if valid else None,
            "kuartal": kuartal(tgl_deal) if valid else None,
            "kontrak_habis": iso(as_date(kol.get(r, "TANGGAL BERAKHIR KONTRAK"))),
            "kontrak_mulai_baru": iso(as_date(kol.get(r, "TANGGAL START KONTRAK NEW"))),
            "kontrak_akhir_baru": iso(as_date(kol.get(r, "TANGGAL END KONTRAK NEW"))),
            "kompensasi": as_num(kol.get(r, "NILAI KOMPENSASI NEW IN RUPIAH")),
            "branding": as_num(kol.get(r, "NILAI BRANDING/ REVISUAL IN RUPIAH")),
            "nilai_total": as_num(kol.get(r, "TOTAL", ke=1)),
            "omset_week": as_num(kol.get(r, "OMSET PER WEEK IN RUPIAH")),
            "ikat_target": upper(kol.get(r, "ADA IKAT TARGET (V)")),
            "valid": valid,
        })
    return deals, issues, mentah


def parse_noo(kol, data, channel_ke_kpi):
    """NOO: baris dengan status DEAL dan/atau sudah terbit nomor AP."""
    deals, issues, mentah = [], [], []
    for r in data:
        afps = upper(kol.get(r, "NAMA AFPS"))
        region = upper(kol.get(r, "REGION"))
        gr = rapikan_grfpm(kol.get(r, "GRFPM"))
        mentah.append((afps, region, gr))

        status = upper(kol.get(r, "DEAL/ PROSES")) or "BELUM DIVISIT"
        if status.startswith("NO DEAL"):
            status = "NO DEAL"
        channel = upper(kol.get(r, "CHANNEL"))
        ap = norm(kol.get(r, "NOMOR AP"))
        if status != "DEAL" and not ap:
            continue

        no_outlet = norm(kol.get(r, "No Outlet"))
        outlet = norm(kol.get(r, "NAMA OUTLET"))
        tgl_deal = as_date(kol.get(r, "TGL DEAL"))
        masalah = []
        if status == "DEAL" and not ap:
            masalah.append("Status DEAL tetapi nomor AP belum terisi (belum diakui sebagai NOO ber-AP)")
        if ap and status != "DEAL":
            masalah.append(f"Sudah ada nomor AP tetapi status masih '{status or 'kosong'}'")
        if tgl_deal is None:
            masalah.append("Sudah DEAL/ber-AP tetapi TGL DEAL kosong")
        elif tgl_deal.year not in TAHUN_VALID:
            masalah.append(f"Tanggal deal di luar rentang wajar: {tgl_deal.isoformat()}")
        for m in masalah:
            issues.append({"sheet": "NOO", "no_outlet": no_outlet, "outlet": outlet,
                           "afps": afps, "region": region, "masalah": m})

        valid = bool(ap) and tgl_deal is not None and tgl_deal.year in TAHUN_VALID
        deals.append({
            "sumber": "NOO",
            "no_outlet": no_outlet,
            "outlet": outlet,
            "grfpm": gr,
            "region": region,
            "kota": upper(kol.get(r, "KOTA")),
            "kecamatan": upper(kol.get(r, "KECAMATAN")),
            "afps": afps,
            "jenis": upper(kol.get(r, "JENIS OUTLET")),
            "channel": channel,
            "kpi": channel_ke_kpi.get(channel, "lainnya"),
            "kategori": upper(kol.get(r, "KATEGORI KPI")),
            "brand": upper(kol.get(r, "BRAND", ke=1)) or upper(kol.get(r, "BRAND")),
            "kontrak": upper(kol.get(r, "BB/BR/BL")),
            "takeover": upper(kol.get(r, "TAKEOVER/BUKAN TAKEOVER")),
            "alamat": norm(kol.get(r, "ALAMAT OUTLET")),
            "pic": norm(kol.get(r, "NAMA PIC")),
            "telp": norm(kol.get(r, "NO TELP PIC")),
            "kode_outlet": norm(kol.get(r, "KODE OUTLET")),
            "ap": ap,
            "spsd": norm(kol.get(r, "NO SPSD")),
            "prioritas": upper(kol.get(r, "PRIORITAS/TIDAK PRIORITAS")),
            "rating": as_num(kol.get(r, "RATING GOOGLE")),
            "skor": as_num(kol.get(r, "TOTAL SCORE (Kolom N x Kolom O)")),
            "siswa": as_int(kol.get(r, "JUMLAH SISWA")),
            "tgl_visit": iso(as_date(kol.get(r, "TGL VISIT"))),
            "tgl_deal": iso(tgl_deal),
            "week_deal": as_int(kol.get(r, "WEEK DEAL")),
            "week_iso": week_iso(tgl_deal) if (tgl_deal and tgl_deal.year in TAHUN_VALID) else None,
            "kuartal": kuartal(tgl_deal) if (tgl_deal and tgl_deal.year in TAHUN_VALID) else None,
            "kontrak_mulai_baru": iso(as_date(kol.get(r, "TANGGAL START KONTRAK BASED ON PKS"))),
            "kontrak_akhir_baru": iso(as_date(kol.get(r, "TANGGAL END KONTRAK BASED ON PKS"))),
            "kompensasi": as_num(kol.get(r, "NILAI KOMPENSASI NEW IN RUPIAH")),
            "branding": as_num(kol.get(r, "NILAI BRANDING/ REVISUAL IN RUPIAH")),
            "nilai_total": as_num(kol.get(r, "TOTAL")),
            "omset_karton": as_num(kol.get(r, "OMSET PER WEEK IN KARTON")),
            "ikat_target": upper(kol.get(r, "ADA IKAT TARGET (V)")),
            "valid": valid,
        })
    return deals, issues, mentah


# ---------------------------------------------------------------- main

def main():
    p = argparse.ArgumentParser(description="Konversi SWS xlsx -> dashboard.json")
    p.add_argument("xlsx", help="Path file SWS (.xlsx)")
    p.add_argument("-o", "--output", default=str(ROOT / "data" / "dashboard.json"))
    p.add_argument("-c", "--config", default=str(ROOT / "config" / "scheme.json"))
    p.add_argument("--week", type=int, default=None, help="Nomor minggu (default: dibaca dari nama file)")
    args = p.parse_args()

    src = Path(args.xlsx)
    if not src.exists():
        raise SystemExit(f"File tidak ditemukan: {src}")

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    channel_ke_kpi = {upper(k): v for k, v in cfg.get("channel_ke_kpi", {}).items()}

    week = args.week
    if week is None:
        m = re.search(r"[_\-\s]?W(\d{1,2})", src.stem, re.IGNORECASE)
        week = int(m.group(1)) if m else None

    print(f"Membaca {src.name} ...", flush=True)
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    tersedia = set(wb.sheetnames)
    for wajib in ("EXT", "NOO"):
        if wajib not in tersedia:
            raise SystemExit(f"Sheet '{wajib}' tidak ada di {src.name}. Sheet tersedia: {sorted(tersedia)}")

    kol_ext, data_ext = baca_sheet(wb, "EXT")
    kol_noo, data_noo = baca_sheet(wb, "NOO")
    print(f"  EXT: {len(data_ext)} baris | NOO: {len(data_noo)} baris", flush=True)

    deals_ext, iss_ext, mentah_ext = parse_ext(kol_ext, data_ext, channel_ke_kpi)
    deals_noo, iss_noo, mentah_noo = parse_noo(kol_noo, data_noo, channel_ke_kpi)

    peta_afps = region_per_afps(mentah_ext + mentah_noo)
    deals = deals_ext + deals_noo
    for d in deals:  # samakan region tiap AFPS supaya agregasi RFPM konsisten
        info = peta_afps.get(d["afps"])
        if info:
            d["region"] = info["region"] or d["region"]
            d["grfpm"] = d["grfpm"] or info["grfpm"]

    per_region = defaultdict(set)
    for afps, info in peta_afps.items():
        per_region[info["region"]].add(afps)

    doc = {
        "meta": {
            "week": week,
            "file_sumber": src.name,
            "dibuat": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
            "grfpm": Counter(d["grfpm"] for d in deals if d["grfpm"]).most_common(1)[0][0] if deals else "",
            "ringkas": {
                "ext_total": len(data_ext),
                "ext_deal": len(deals_ext),
                "noo_total": len(data_noo),
                "noo_deal_ap": len(deals_noo),
                "jumlah_afps": len(peta_afps),
                "jumlah_region": len(per_region),
            },
        },
        "afps": sorted(
            ({"nama": a, "region": i["region"], "grfpm": i["grfpm"]} for a, i in peta_afps.items()),
            key=lambda x: (x["region"], x["nama"]),
        ),
        "region": sorted(
            ({"nama": r, "afps": sorted(a), "jumlah_afps": len(a)} for r, a in per_region.items()),
            key=lambda x: x["nama"],
        ),
        "deals": deals,
        "issues": iss_ext + iss_noo,
    }

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(doc, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    print(f"  Deal perpanjangan : {len(deals_ext)}")
    print(f"  NOO ber-AP        : {len(deals_noo)}")
    print(f"  Catatan data      : {len(doc['issues'])}")
    print(f"  AFPS / Region     : {len(peta_afps)} / {len(per_region)}")
    print(f"Ditulis ke {out} ({out.stat().st_size / 1024:.0f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
